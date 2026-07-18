#!/usr/bin/env python3
"""
Gerador de Relatório Excel
==========================
Transforma um CSV de vendas em um relatório .xlsx formatado, com aba de dados,
aba de resumo com KPIs calculados por fórmula, tabelas dinâmicas de apoio e
gráficos (barras por região e linha por mês).

Esquema esperado do CSV:
    data, regiao, produto, categoria, quantidade, preco_unitario, receita

Uso:
    python gerador_relatorio_excel.py --gerar-exemplo exemplo_vendas.csv
    python gerador_relatorio_excel.py --input exemplo_vendas.csv --output relatorio_vendas.xlsx
"""

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

COLUNAS_ESPERADAS = ["data", "regiao", "produto", "categoria", "quantidade", "preco_unitario", "receita"]
FONTE = "Arial"
MOEDA = "R$ #,##0.00"
COR_CABECALHO = "1F4E79"


def gerar_exemplo(caminho: Path, linhas: int = 500, seed: int = 42) -> None:
    """Gera um CSV de vendas fictícias para demonstração."""
    rng = np.random.default_rng(seed)
    produtos = {
        "Notebook": ("Eletrônicos", 3500.0), "Monitor": ("Eletrônicos", 1200.0),
        "Teclado": ("Periféricos", 250.0), "Mouse": ("Periféricos", 120.0),
        "Cadeira Gamer": ("Móveis", 1500.0), "Mesa": ("Móveis", 900.0),
    }
    nomes = rng.choice(list(produtos), size=linhas)
    quantidade = rng.integers(1, 8, size=linhas)
    preco = np.array([produtos[p][1] for p in nomes]) * rng.uniform(0.9, 1.1, size=linhas)

    df = pd.DataFrame({
        "data": pd.to_datetime("2025-01-01") + pd.to_timedelta(rng.integers(0, 365, size=linhas), unit="D"),
        "regiao": rng.choice(["Sudeste", "Sul", "Nordeste", "Centro-Oeste", "Norte"],
                             size=linhas, p=[0.4, 0.2, 0.2, 0.1, 0.1]),
        "produto": nomes,
        "categoria": [produtos[p][0] for p in nomes],
        "quantidade": quantidade,
        "preco_unitario": preco.round(2),
    })
    df["receita"] = (df["quantidade"] * df["preco_unitario"]).round(2)
    df.sort_values("data").to_csv(caminho, index=False)
    print(f"[OK] Exemplo gerado: {caminho} ({linhas} linhas)")


def _estilizar_cabecalho(ws, n_colunas: int, linha: int = 1) -> None:
    fill = PatternFill("solid", fgColor=COR_CABECALHO)
    for col in range(1, n_colunas + 1):
        celula = ws.cell(row=linha, column=col)
        celula.font = Font(name=FONTE, bold=True, color="FFFFFF")
        celula.fill = fill
        celula.alignment = Alignment(horizontal="center")


def _ajustar_larguras(ws) -> None:
    for col_cells in ws.columns:
        letra = get_column_letter(col_cells[0].column)
        largura = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[letra].width = min(largura + 3, 40)


def montar_relatorio(df: pd.DataFrame, saida: Path) -> None:
    """Escreve o relatório .xlsx com abas Dados e Resumo."""
    df = df.copy()
    df["data"] = pd.to_datetime(df["data"])
    n = len(df)

    wb = Workbook()

    # ---- Aba Dados -----------------------------------------------------
    ws_dados = wb.active
    ws_dados.title = "Dados"
    for linha in dataframe_to_rows(df.assign(data=df["data"].dt.strftime("%d/%m/%Y")), index=False, header=True):
        ws_dados.append(linha)
    for row in ws_dados.iter_rows():
        for c in row:
            c.font = Font(name=FONTE)
    for col in ("F", "G"):  # preco_unitario e receita
        for c in ws_dados[col][1:]:
            c.number_format = MOEDA
    _estilizar_cabecalho(ws_dados, df.shape[1])
    ws_dados.freeze_panes = "A2"
    ws_dados.auto_filter.ref = ws_dados.dimensions
    _ajustar_larguras(ws_dados)

    # ---- Aba Resumo: KPIs por fórmula ----------------------------------
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "Relatório de Vendas — Resumo Executivo"
    ws["A1"].font = Font(name=FONTE, size=14, bold=True)

    kpis = [
        ("Receita total", f"=SUM(Dados!$G$2:$G${n + 1})", MOEDA),
        ("Pedidos", f"=COUNTA(Dados!$A$2:$A${n + 1})", "#,##0"),
        ("Itens vendidos", f"=SUM(Dados!$E$2:$E${n + 1})", "#,##0"),
        ("Ticket médio", f"=AVERAGE(Dados!$G$2:$G${n + 1})", MOEDA),
    ]
    for i, (rotulo, formula, fmt) in enumerate(kpis, start=3):
        ws.cell(row=i, column=1, value=rotulo).font = Font(name=FONTE, bold=True)
        celula = ws.cell(row=i, column=2, value=formula)
        celula.font = Font(name=FONTE)
        celula.number_format = fmt

    # ---- Tabelas de apoio (pandas) + gráficos --------------------------
    por_regiao = df.groupby("regiao", as_index=False)["receita"].sum().sort_values("receita", ascending=False)
    por_mes = (df.assign(mes=df["data"].dt.strftime("%Y-%m"))
                 .groupby("mes", as_index=False)["receita"].sum())

    def escrever_tabela(tabela: pd.DataFrame, col_inicial: int, titulo: str) -> tuple[int, int]:
        ws.cell(row=2, column=col_inicial, value=titulo).font = Font(name=FONTE, bold=True)
        for i, linha in enumerate([list(tabela.columns)] + tabela.values.tolist(), start=3):
            for j, valor in enumerate(linha):
                c = ws.cell(row=i, column=col_inicial + j, value=valor)
                c.font = Font(name=FONTE)
                if i > 3 and j == 1:
                    c.number_format = MOEDA
        for j in range(2):
            ws.cell(row=3, column=col_inicial + j).font = Font(name=FONTE, bold=True)
        return 3, 3 + len(tabela)

    inicio_r, fim_r = escrever_tabela(por_regiao, 4, "Receita por região")
    inicio_m, fim_m = escrever_tabela(por_mes, 7, "Receita por mês")

    grafico_barras = BarChart()
    grafico_barras.title = "Receita por região"
    grafico_barras.add_data(Reference(ws, min_col=5, min_row=inicio_r, max_row=fim_r), titles_from_data=True)
    grafico_barras.set_categories(Reference(ws, min_col=4, min_row=inicio_r + 1, max_row=fim_r))
    grafico_barras.height, grafico_barras.width = 8, 14
    ws.add_chart(grafico_barras, "A10")

    grafico_linha = LineChart()
    grafico_linha.title = "Evolução mensal da receita"
    grafico_linha.add_data(Reference(ws, min_col=8, min_row=inicio_m, max_row=fim_m), titles_from_data=True)
    grafico_linha.set_categories(Reference(ws, min_col=7, min_row=inicio_m + 1, max_row=fim_m))
    grafico_linha.height, grafico_linha.width = 8, 14
    ws.add_chart(grafico_linha, "A28")

    _ajustar_larguras(ws)
    wb.save(saida)
    print(f"[OK] Relatório gerado: {saida}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera relatório Excel formatado a partir de um CSV de vendas.")
    parser.add_argument("--input", type=Path, help="CSV de vendas de entrada")
    parser.add_argument("--output", type=Path, default=Path("relatorio_vendas.xlsx"), help="Arquivo .xlsx de saída")
    parser.add_argument("--gerar-exemplo", type=Path, metavar="CSV", help="Gera um CSV de exemplo e encerra")
    args = parser.parse_args()

    if args.gerar_exemplo:
        gerar_exemplo(args.gerar_exemplo)
        return
    if not args.input:
        parser.error("informe --input (ou use --gerar-exemplo para criar dados de teste)")
    if not args.input.exists():
        print(f"[ERRO] Arquivo não encontrado: {args.input}", file=sys.stderr)
        sys.exit(1)

    df = pd.read_csv(args.input)
    faltantes = [c for c in COLUNAS_ESPERADAS if c not in df.columns]
    if faltantes:
        print(f"[ERRO] Colunas ausentes no CSV: {faltantes}", file=sys.stderr)
        sys.exit(1)

    montar_relatorio(df[COLUNAS_ESPERADAS], args.output)


if __name__ == "__main__":
    main()
